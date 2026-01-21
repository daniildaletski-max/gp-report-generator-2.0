#!/usr/bin/env python3
"""
Generate Excel file with native chart based on FM-MonthlyTeamOverviewTemplate.xlsx
This script creates a proper Excel chart that works in both Excel and Google Sheets.
"""

import sys
import json
import openpyxl
from openpyxl.chart import BarChart, Reference, Series
from openpyxl.chart.series import DataPoint
from openpyxl.chart.label import DataLabelList
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from copy import copy
import os

def create_report_with_chart(data_json_or_path, output_path, template_path=None):
    """
    Create Excel report with embedded chart.
    
    Args:
        data_json_or_path: JSON string with report data OR path to JSON file
        output_path: Path to save the output Excel file
        template_path: Optional path to template file
    """
    # Check if it's a file path or JSON string
    if os.path.isfile(data_json_or_path):
        with open(data_json_or_path, 'r', encoding='utf-8') as f:
            data = json.load(f)
    else:
        data = json.loads(data_json_or_path)
    
    # Create workbook
    if template_path and os.path.exists(template_path):
        wb = openpyxl.load_workbook(template_path)
        ws = wb.active
    else:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "TEMPLATE"
    
    # Extract data
    team_name = data.get('teamName', 'Unknown Team')
    month_name = data.get('monthName', 'Unknown Month')
    year = data.get('year', 2026)
    gp_data = data.get('gpData', [])
    fm_performance = data.get('fmPerformance', '')
    goals_this_month = data.get('goalsThisMonth', '')
    team_overview = data.get('teamOverview', '')
    
    # Colors matching the template
    purple_main = PatternFill(start_color="C3BBFD", end_color="C3BBFD", fill_type="solid")
    purple_light = PatternFill(start_color="C5C8FB", end_color="C5C8FB", fill_type="solid")
    purple_lighter = PatternFill(start_color="DAD8FE", end_color="DAD8FE", fill_type="solid")
    green_fill = PatternFill(start_color="92D050", end_color="92D050", fill_type="solid")
    yellow_fill = PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid")
    red_fill = PatternFill(start_color="FF6B6B", end_color="FF6B6B", fill_type="solid")
    
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # ===== HEADER SECTION =====
    # Team Name (C2)
    ws.merge_cells('C2:L2')
    ws['C2'] = team_name
    ws['C2'].font = Font(bold=True, size=14)
    ws['C2'].fill = purple_main
    ws['C2'].alignment = Alignment(horizontal='center', vertical='center')
    
    # Month Overview (N2)
    ws.merge_cells('N2:X2')
    ws['N2'] = f"{month_name} Overview"
    ws['N2'].font = Font(bold=True, size=14)
    ws['N2'].fill = purple_main
    ws['N2'].alignment = Alignment(horizontal='center', vertical='center')
    
    # ===== FM PERFORMANCE SECTION (B3:H19) =====
    ws.merge_cells('B3:H3')
    ws['B3'] = "FM Performance"
    ws['B3'].font = Font(bold=True)
    ws['B3'].fill = purple_light
    
    ws.merge_cells('B4:H19')
    ws['B4'] = fm_performance
    ws['B4'].alignment = Alignment(wrap_text=True, vertical='top')
    
    # ===== ATTENDANCE TABLE HEADERS (N3:X3) =====
    # Header structure: Name (N3:P3), Score (Q3), Mistakes (R3), Extra shifts (S3), Lateness (T3), Missed day (U3), Sick leave (V3), Attitude (W3), Remarks (X3)
    
    # Name header (merged N3:P3)
    ws.merge_cells('N3:P3')
    ws['N3'] = 'Name'
    ws['N3'].font = Font(bold=True, size=9)
    ws['N3'].fill = purple_light
    ws['N3'].border = thin_border
    ws['N3'].alignment = Alignment(horizontal='center', vertical='center')
    
    # Individual headers for other columns (Q onwards)
    attendance_headers = [
        ('Q3', ''),  # Score column (empty header, data below)
        ('R3', 'Mistakes'),
        ('S3', 'Extra shifts'),
        ('T3', 'Lateness'),
        ('U3', 'Missed day'),
        ('V3', 'Sick leave'),
        ('W3', 'Attitude'),
        ('X3', 'Remarks')
    ]
    
    for cell_ref, header in attendance_headers:
        cell = ws[cell_ref]
        cell.value = header
        cell.font = Font(bold=True, size=9)
        cell.fill = purple_light
        cell.border = thin_border
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    
    # ===== GP DATA ROWS (N4:X33, 2 rows per GP) =====
    row = 4
    chart_data_names = []
    chart_data_scores = []
    chart_data_prev = []
    
    for i, gp in enumerate(gp_data[:15]):  # Max 15 GPs
        gp_name = gp.get('name', f'GP {i+1}')
        score = gp.get('score', 0)
        prev_score = gp.get('prevScore', 0)
        mistakes = gp.get('mistakes', 0)
        extra_shifts = gp.get('extraShifts', 0)
        lateness = gp.get('lateness', 0)
        missed_days = gp.get('missedDays', 0)
        sick_leave = gp.get('sickLeave', 0)
        attitude = gp.get('attitude', '')
        remarks = gp.get('remarks', '')
        
        chart_data_names.append(gp_name)
        chart_data_scores.append(score)
        chart_data_prev.append(prev_score)
        
        # Merge cells for GP name (N:P, 2 rows)
        ws.merge_cells(f'N{row}:P{row+1}')
        ws[f'N{row}'] = gp_name
        ws[f'N{row}'].alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        ws[f'N{row}'].border = thin_border
        
        # Score cell (Q only, 2 rows)
        ws.merge_cells(f'Q{row}:Q{row+1}')
        ws[f'Q{row}'] = score
        ws[f'Q{row}'].alignment = Alignment(horizontal='center', vertical='center')
        ws[f'Q{row}'].font = Font(bold=True, size=12)
        ws[f'Q{row}'].border = thin_border
        
        # Color based on score
        if score >= 18:
            ws[f'Q{row}'].fill = green_fill
        elif score >= 15:
            ws[f'Q{row}'].fill = yellow_fill
        elif score > 0:
            ws[f'Q{row}'].fill = red_fill
        
        # Other columns (R:X) - Mistakes, Extra shifts, Lateness, Missed day, Sick leave, Attitude, Remarks
        other_data = [mistakes, extra_shifts, lateness, missed_days, sick_leave, attitude, remarks]
        for j, val in enumerate(other_data):
            col = 18 + j  # Column R onwards (R=18)
            ws.merge_cells(f'{get_column_letter(col)}{row}:{get_column_letter(col)}{row+1}')
            ws.cell(row=row, column=col).value = val if val else ''
            ws.cell(row=row, column=col).alignment = Alignment(horizontal='center', vertical='center')
            ws.cell(row=row, column=col).border = thin_border
        
        row += 2
    
    # ===== TEAM MANAGEMENT SECTION (Row 23) =====
    ws.merge_cells('B23:H23')
    ws['B23'] = "Team Management"
    ws['B23'].font = Font(bold=True)
    ws['B23'].fill = purple_main
    
    # Goals this month (B24:E39)
    ws.merge_cells('B24:E24')
    ws['B24'] = "Goals this month"
    ws['B24'].font = Font(bold=True)
    ws['B24'].fill = purple_light
    
    ws.merge_cells('B25:E39')
    ws['B25'] = goals_this_month
    ws['B25'].alignment = Alignment(wrap_text=True, vertical='top')
    
    # Team Overview (F24:H39)
    ws.merge_cells('F24:H24')
    ws['F24'] = "Team Overview"
    ws['F24'].font = Font(bold=True)
    ws['F24'].fill = purple_light
    
    ws.merge_cells('F25:H39')
    ws['F25'] = team_overview
    ws['F25'].alignment = Alignment(wrap_text=True, vertical='top')
    
    # ===== CHART DATA (Hidden area for chart reference) =====
    chart_data_start_row = 60
    ws.cell(row=chart_data_start_row, column=14).value = "GP Name"
    ws.cell(row=chart_data_start_row, column=15).value = f"{month_name} {year}"
    ws.cell(row=chart_data_start_row, column=16).value = "Previous Month"
    
    for i, (name, score, prev) in enumerate(zip(chart_data_names, chart_data_scores, chart_data_prev)):
        ws.cell(row=chart_data_start_row + 1 + i, column=14).value = name
        ws.cell(row=chart_data_start_row + 1 + i, column=15).value = score
        ws.cell(row=chart_data_start_row + 1 + i, column=16).value = prev
    
    chart_data_end_row = chart_data_start_row + len(chart_data_names)
    
    # Note: Chart data rows are visible but placed below the main content
    # They can be hidden manually in Google Sheets if needed
    
    # ===== CREATE CHART =====
    if len(chart_data_names) > 0:
        chart = BarChart()
        chart.type = "col"
        chart.grouping = "clustered"
        chart.style = 10
        chart.title = f"{team_name} - {month_name} {year} GP Performance"
        chart.y_axis.title = "Score"
        chart.x_axis.title = "Game Presenter"
        
        # Data references
        data_ref = Reference(ws, min_col=15, min_row=chart_data_start_row, 
                        max_row=chart_data_end_row, max_col=16)
        cats_ref = Reference(ws, min_col=14, min_row=chart_data_start_row + 1, 
                        max_row=chart_data_end_row)
        
        chart.add_data(data_ref, titles_from_data=True)
        chart.set_categories(cats_ref)
        chart.shape = 4
        
        # Set chart size
        chart.width = 18
        chart.height = 10
        
        # Position chart below attendance table (around N36)
        ws.add_chart(chart, "N36")
    
    # ===== ADDITIONAL NOTES SECTION (Row 42) =====
    ws.merge_cells('B42:H42')
    ws['B42'] = "Additional Notes"
    ws['B42'].font = Font(bold=True)
    ws['B42'].fill = purple_main
    
    ws.merge_cells('B43:H58')
    ws['B43'] = data.get('additionalNotes', '')
    ws['B43'].alignment = Alignment(wrap_text=True, vertical='top')
    
    # ===== SET COLUMN WIDTHS =====
    column_widths = {
        'A': 3, 'B': 12, 'C': 12, 'D': 12, 'E': 12, 'F': 12, 'G': 12, 'H': 12,
        'I': 3, 'J': 3, 'K': 3, 'L': 3, 'M': 3,
        'N': 12, 'O': 6, 'P': 6, 'Q': 8, 'R': 8, 'S': 8, 'T': 8, 'U': 8, 'V': 8, 'W': 8, 'X': 12
    }
    for col, width in column_widths.items():
        ws.column_dimensions[col].width = width
    
    # Save workbook
    wb.save(output_path)
    print(f"SUCCESS: {output_path}")
    return output_path


if __name__ == "__main__":
    if len(sys.argv) < 3:
        print("Usage: python generate_excel_chart.py <json_data> <output_path> [template_path]")
        sys.exit(1)
    
    data_json = sys.argv[1]
    output_path = sys.argv[2]
    template_path = sys.argv[3] if len(sys.argv) > 3 else None
    
    try:
        result = create_report_with_chart(data_json, output_path, template_path)
        print(result)
    except Exception as e:
        print(f"ERROR: {str(e)}", file=sys.stderr)
        sys.exit(1)
